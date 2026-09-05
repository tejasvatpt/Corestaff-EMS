"""Seed 50 realistic employees via the Admin Onboarding endpoint and export
their temporary credentials to an Excel (.xlsx) file.
"""

import os
import sys
import random
from datetime import date, timedelta
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure project root is in sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.auth import create_access_token
from app.database import SessionLocal
from app.models import Department, User

API_BASE = "http://127.0.0.1:8000"

FIRST_NAMES = [
    "Aarav", "Priya", "Alexander", "Sophia", "Liam", "Emma", "Noah", "Olivia",
    "Rohan", "Ananya", "Ethan", "Ava", "Lucas", "Isabella", "Aditya", "Ishita",
    "Mason", "Mia", "Oliver", "Harper", "Kabir", "Diya", "Elijah", "Evelyn",
    "James", "Abigail", "Benjamin", "Emily", "Dev", "Neha", "Sebastian", "Ella",
    "Jack", "Scarlett", "Arjun", "Kavya", "Daniel", "Chloe", "Henry", "Victoria",
    "Siddharth", "Tanvi", "Owen", "Grace", "Jackson", "Zoey", "Vihaan", "Meera",
    "Mateo", "Penelope", "Vivaan", "Riya", "Aiden", "Maya", "Leo", "Layla",
    "Lucas", "Zara", "Julian", "Aria"
]

LAST_NAMES = [
    "Sharma", "Patel", "Smith", "Johnson", "Williams", "Brown", "Jones", "Miller",
    "Verma", "Gupta", "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
    "Iyer", "Rao", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson",
    "Reddy", "Nair", "Martin", "Lee", "Perez", "Thompson", "White", "Harris",
    "Chopra", "Malhotra", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson", "Walker",
    "Kapoor", "Bhatia", "Young", "Allen", "King", "Wright", "Scott", "Torres",
    "Mehta", "Saxena", "Das", "Joshi", "Bansal", "Chaudhary"
]

DESIGNATIONS = {
    "Engineering": [
        "Senior Backend Engineer", "Frontend Developer", "Full Stack Engineer",
        "DevOps & Cloud Engineer", "QA Automation Engineer", "Tech Lead",
        "Mobile App Developer", "Systems Architect"
    ],
    "Product & Design": [
        "Lead Product Manager", "UI/UX Designer", "Product Designer",
        "Technical Product Manager", "Design Systems Lead"
    ],
    "Marketing": [
        "Growth Marketing Lead", "Content Strategist", "SEO Specialist",
        "Brand Manager", "Performance Marketer"
    ],
    "Human Resources": [
        "HR Business Partner", "Talent Acquisition Specialist", "People Operations Lead",
        "Compensation & Benefits Specialist"
    ],
    "Finance & Operations": [
        "Financial Analyst", "Operations Manager", "Accounting Lead",
        "Billing Specialist"
    ],
    "Customer Success": [
        "Customer Success Manager", "Support Operations Lead", "Client Solutions Engineer"
    ],
}


def run_onboarding_seed(target_count=50):
    print(f"Starting Admin-Led Onboarding of {target_count} employees...")

    # 1. Obtain Admin JWT Token
    admin_token = create_access_token({"sub": "tejasva.admin@gmail.com"})
    headers = {
        "Authorization": f"Bearer {admin_token}",
        "Content-Type": "application/json",
    }

    # 2. Query departments
    db = SessionLocal()
    departments = db.query(Department).all()
    if not departments:
        # Fallback create standard departments
        for dept_name in DESIGNATIONS.keys():
            db.add(Department(name=dept_name))
        db.commit()
        departments = db.query(Department).all()

    dept_map = {d.name: d.id for d in departments}
    dept_list = list(departments)
    print(f"Available Departments: {[d.name for d in dept_list]}")

    # 3. Track existing emails
    existing_emails = set(row[0] for row in db.query(User.email).all())
    db.close()

    # 4. Generate pairs
    random.seed(42)  # For reproducible realistic data
    pairs = []
    for fn in FIRST_NAMES:
        for ln in LAST_NAMES:
            pairs.append((fn, ln))
    random.shuffle(pairs)

    onboarded_records = []
    pair_idx = 0
    today = date.today()

    while len(onboarded_records) < target_count and pair_idx < len(pairs):
        fn, ln = pairs[pair_idx]
        pair_idx += 1

        full_name = f"{fn} {ln}"
        suffix = random.randint(100, 999)
        email = f"{fn.lower()}.{ln.lower()}{suffix}@company.corp"

        if email in existing_emails:
            continue

        chosen_dept = random.choice(dept_list)
        dept_id = chosen_dept.id
        dept_name = chosen_dept.name

        desigs = DESIGNATIONS.get(dept_name, ["Team Member", "Specialist", "Associate"])
        designation = random.choice(desigs)

        phone = f"+1 ({random.randint(200, 999)}) {random.randint(100, 999)}-{random.randint(1000, 9999)}"
        days_ago = random.randint(10, 500)
        joining_date = (today - timedelta(days=days_ago)).isoformat()

        payload = {
            "full_name": full_name,
            "email": email,
            "department_id": dept_id,
            "designation": designation,
            "phone": phone,
            "joining_date": joining_date,
        }

        try:
            res = requests.post(f"{API_BASE}/employees/onboard", json=payload, headers=headers)
            if res.status_code == 200:
                data = res.json()
                onboarded_records.append({
                    "id": data["id"],
                    "full_name": data["full_name"],
                    "email": data["email"],
                    "username": data["username"],
                    "temp_password": data["temp_password"],
                    "department": data["department_name"] or dept_name,
                    "designation": data["designation"] or designation,
                    "phone": data["phone"] or phone,
                    "joining_date": data["joining_date"] or joining_date,
                })
                existing_emails.add(email)
                print(f"[{len(onboarded_records)}/{target_count}] Onboarded: {full_name} ({email})")
            else:
                print(f"Failed to onboard {full_name}: {res.status_code} - {res.text}")
        except Exception as e:
            print(f"Error calling onboard API: {e}")
            break

    print(f"\nSuccessfully onboarded {len(onboarded_records)} employees through Admin API!")

    # 5. Export to Excel
    export_to_excel(onboarded_records)


def export_to_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Credentials"

    # Styling constants
    header_fill = PatternFill(start_color="1D1D1F", end_color="1D1D1F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
    code_font = Font(name="Consolas", size=10, color="1D1D1F")
    regular_font = Font(name="Calibri", size=10, color="1D1D1F")

    thin_border = Border(
        left=Side(style="thin", color="D2D2D7"),
        right=Side(style="thin", color="D2D2D7"),
        top=Side(style="thin", color="D2D2D7"),
        bottom=Side(style="thin", color="D2D2D7")
    )

    headers = [
        "Employee ID",
        "Full Name",
        "Corporate Email (Login ID)",
        "Username",
        "Temporary Password",
        "Department",
        "Designation",
        "Phone Number",
        "Joining Date",
        "Status"
    ]

    ws.append(headers)

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
    ws.row_dimensions[1].height = 26

    # Append data rows
    for row_idx, r in enumerate(records, start=2):
        row_data = [
            r["id"],
            r["full_name"],
            r["email"],
            r["username"],
            r["temp_password"],
            r["department"],
            r["designation"],
            r["phone"],
            r["joining_date"],
            "Active"
        ]
        ws.append(row_data)

        use_zebra = (row_idx % 2 == 0)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            if col_idx in (3, 4, 5):  # email, username, temp_password
                cell.font = code_font
            else:
                cell.font = regular_font

            if use_zebra:
                cell.fill = zebra_fill

            # Alignment
            if col_idx in (1, 9, 10):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_idx].height = 20

    # Auto-fit column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to both workspace root and scripts/ folder for easy access
    paths = [
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "employee_credentials.xlsx"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "employee_credentials.xlsx"),
    ]

    for path in paths:
        wb.save(path)
        print(f"Credentials Excel file saved to: {os.path.abspath(path)}")


if __name__ == "__main__":
    run_onboarding_seed(50)
